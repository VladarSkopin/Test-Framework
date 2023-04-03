from typing import Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_max_row(file: str, sheet_name: str) -> int:
    workbook: Workbook = openpyxl.load_workbook(file)
    sheet: Worksheet = workbook[sheet_name]
    return sheet.max_row


def get_max_column(file: str, sheet_name: str) -> int:
    workbook: Workbook = openpyxl.load_workbook(file)
    sheet: Worksheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(file: str, sheet_name: str, row_number: int, column_number: int) -> Optional[str | int | float]:
    workbook: Workbook = openpyxl.load_workbook(file)
    sheet: Worksheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def write_data(file: str, sheet_name: str, row_number: int, column_number: int, data: Optional[str | int | float]) -> \
        None:
    workbook: Workbook = openpyxl.load_workbook(file)
    sheet: Worksheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(file)
